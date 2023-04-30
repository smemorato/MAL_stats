import argparse
import json
from mal_token import refresh_token, get_new_code_verifier,print_new_authorisation_url,\
    generate_new_token,print_user_info
import mal_api
from openpyxl import load_workbook
import comparison_excel


if __name__ == "__main__":
    parser = argparse.ArgumentParser(     prog='PROG',
        description='''Create an excel file to compare user stats''',
        epilog='''
        Hope  it works!!!''')

    parser.add_argument("-fn", "--filename",
                        required=False,
                        help="pick the name of the file",
                        metavar=" ")

    args = parser.parse_args()



    file = "user comparison.xlsx"
    workbook = load_workbook(filename=file)
    sheet = workbook['users']
    row = 2
    lists = []
    #print(sheet.tables)
    users = sheet[sheet.tables["users_table"].ref]
    content = []
    for ent in users:
        content.append(ent[1].value)

    # content2 = [[cell.value for cell in ent]
    #            for ent in users
    #            ]

    if "mal" in content:
        try:
            with open('token.json', 'r') as file:
                print(file)
                token = json.load(file)
                rtoken = refresh_token(token)

        except Exception as err:
            print(err)
            # todo make this a function in mal_api
            code_verifier = code_challenge = get_new_code_verifier()
            print_new_authorisation_url(code_challenge)
            authorisation_code = input('Copy-paste the Authorisation Code: ').strip()
            token = generate_new_token(authorisation_code, code_verifier)
            print_user_info(token['access_token'])

    while sheet.cell(row=row, column=1).value is not None:

        if sheet.cell(row=row, column=2).value == "mal":
            print(sheet.cell(row=row, column=1).value)
            response = mal_api.request_list(sheet.cell(row=row, column=1).value, rtoken["access_token"])
            if response[1] == 200:

                lists.append([sheet.cell(row=row, column=1).value,  response[0]])
                #print(sheet.cell(row=row, column=1).value)
            elif response[1] == 404:
                print("username {} doesn't exist for the given source".format(sheet.cell(row=row, column=1).value))
                print("check whether source or username are correct")
            else:
                print("t failed try again later, server or internet connection may be down")



        else:
            print("anilist")
            print(sheet.cell(row=row, column=1).value)
        row = row + 1

    comparison_excel.to_excel(workbook, lists)
        #comparison_excel.to_excel(response[0], args.username)



