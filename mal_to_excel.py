from openpyxl import load_workbook
import datetime
import json
import pandas as pd
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, DataBarRule, IconSetRule
import numpy as np


def insert_table(df, df_genrestable,df_days, workbook, username,source):

    #if there're more test value on the template file user's list delete teste values
    n = len(df.index)+1
    while workbook['user_list'].cell(row=n, column=2).value is not None:
        workbook['user_list'].delete_rows(n, 1)
        n=n+1

    merged = pd.merge(df, df_genrestable, how="inner", on="ID")

    #if there're more test value on the template file user's list delete teste values
    n = len(merged.index)+1
    while workbook['genres_table'].cell(row=n, column=2).value is not None:
        workbook['genres_table'].delete_rows(n, 1)
        n=n+1


    #this doesn't work with the newer version of pandas/openpyxl
    with pd.ExcelWriter(f"userlist/{username}_{source}.xlsx", engine='openpyxl') as writer:

        # adds workbook and sheets to writer
        writer.book= workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)


        df.to_excel(writer, sheet_name='user_list', header=False, startrow=1, index=False)
        merged.to_excel(writer, sheet_name='genres_table', header=False, startrow=1, index=False)
        df_days.to_excel(writer, sheet_name="days", header=False, startrow=1, index=False)

        writer.sheets["user_list"].tables['tb_list'].ref = 'A1:P' + str(len(df) + 1)
        writer.sheets["genres_table"].tables['tb_anime_genres'].ref = 'A1:Q' + str(len(merged) + 1)
        writer.sheets["days"].tables['tb_days'].ref = 'A1:G' + str(len(df_days) + 1)

        #writer.sheets["progression"].column_dimensions["N"].number_format ="0.00"
        #round numbers
        for r in range(2,len(df_days)+3):
            writer.sheets["days"][f'B{r}'].number_format = "0.00"
            writer.sheets["days"][f'C{r}'].number_format = "0.00"
            writer.sheets["days"][f'D{r}'].number_format = "0.00"




def to_excel(userlist, username):
    file = "user_stats_template.xlsx"
    workbook = load_workbook(filename=file)

    m = 1
    df = pd.DataFrame()
    for page in userlist:
        page_df = pd.json_normalize(page['data'])

        #if user doen't have start or finish date the api doesn't redturn those field so i have to add them
        if {'list_status.start_date', 'list_status.finish_date'}.issubset(page_df.columns):
            page_df.columns = ["ID", "Title", "pic big", "pic medium", "start Date", "End Date", "mean", "type", "genres",
                            "num ep", "season year", "season", "source", "ep duration", "studios", "status", "score",
                            "ep watched", "rewatched", "update", "start date", "finish date"]
        elif {'list_status.start_date'}.issubset(page_df.columns):
            page_df.columns = ["ID", "Title", "pic big", "pic medium", "start Date", "End Date", "mean", "type", "genres",
                "num ep", "season year", "season", "source", "ep duration", "studios", "status", "score",
                "ep watched", "rewatched", "update", "start date"]
            page_df.columns["finish date"]= ""
        elif {'list_status.finish_date'}.issubset(page_df.columns):
            page_df.columns = ["ID", "Title", "pic big", "pic medium", "start Date", "End Date", "mean", "type", "genres",
                "num ep", "season year", "season", "source", "ep duration", "studios", "status", "score",
                "ep watched", "rewatched", "update", "finish date"]
            page_df.columns["start start"]= ""
            page_df[["ID", "Title", "pic big", "pic medium", "start Date", "End Date", "mean", "type", "genres",
                "num ep", "season year", "season", "source", "ep duration", "studios", "status", "score",
                "ep watched", "rewatched", "update","start date", "finish date"]]
        else:
            page_df.columns = ["ID", "Title", "pic big", "pic medium", "start Date", "End Date", "mean", "type", "genres",
                "num ep", "season year", "season", "source", "ep duration", "studios", "status", "score",
                "ep watched", "rewatched", "update"]
            page_df.columns["start start"]= ""
            page_df.columns["finish start"]= ""

        df = pd.concat([df, page_df])


    pd.set_option("display.max_columns", None)

    #dataframe with list of different genres
    df_genreslist = df["genres"].explode().apply(pd.Series)
    df_genreslist["name"].fillna("no Genre", inplace=True)
    df_genreslist = df_genreslist["name"].unique()


    #genre relations dataframe
    df_genrestable = df[["ID", "genres"]].copy()
    df_genrestable = df_genrestable.explode("genres")
    df_genrestable = pd.concat([df_genrestable.drop(["genres"], axis=1), df_genrestable["genres"].apply(pd.Series)], axis=1)
    
    #if there is an anime without any genres in the user list the concat will genrerate an column 0
    if 0 in df.columns:
        df_genrestable.drop(columns=[0], inplace=True)
    df_genrestable.drop(columns=["id"], inplace=True)
    df_genrestable["name"].fillna("no Genre", inplace=True)

    #filter wierd mal dates like 30/2/2022 (30th february)
    df['start Date'] = pd.to_datetime(df['start Date'], errors="coerce")
    df['start date'] = pd.to_datetime(df['start date'], errors="coerce")
    df['finish date'] = pd.to_datetime(df['finish date'], errors="coerce")
    # change episodes duration to hours
    df["ep duration"] = df["ep duration"]/3600

    #add year and season column
    df["aux"] = pd.to_numeric(pd.DatetimeIndex(df['start Date']).month, errors="coerce")
    df["aux"] = df["aux"].replace([[1, 2, 3], [4, 5, 6], [7, 8, 9], [10, 11, 12]], ["winter", "spring", "summer", "fall"])
    df["season"].fillna(df['aux'], inplace=True)
    df['aux'] = pd.to_numeric(pd.DatetimeIndex(df['start Date']).year, errors="coerce")
    df["season year"].fillna(df['aux'], inplace=True)

    df.drop(columns=["pic big", "pic medium", "start Date", "End Date"
        , "genres", "num ep", "studios", "status",
                     "rewatched", "update", "aux"])
    df = df[["ID", "Title", "type", "source", "mean", "start date", "finish date", "score", "season year", "season",
             "ep watched", "ep duration"]]
    df["show duration"] = df["ep duration"]*df["ep watched"]

    #I had this done in spreadsheet but when i resize the table the formula would fill the table so i add this here
    df["days watching"] = (pd.to_datetime(df["finish date"], errors= "coerce") -
                           pd.to_datetime(df["start date"], errors="coerce")) / np.timedelta64(1, 'D')+1
    df["hours a day"] = pd.to_numeric(df["show duration"], errors= "coerce")/df["days watching"]
    df["episodes a day"] = df["ep watched"] / df["days watching"]


    oldest = pd.to_datetime(df['start date'], errors="coerce").min().year

    # get dataframe with date from the oldest date to today
    df_days = resize_days_table(oldest, workbook)

    # resize the table to fit the new entries
    workbook = resize_table_progression(workbook, oldest)

    workbook = resize_genres(workbook, df_genreslist)


    #fill forecast table
    #TODO finish this
    workbook = forecast_table(workbook,oldest)

    # add genres to table
    insert_table(df, df_genrestable, df_days, workbook, username, "mal")



#table with list of dates from the oldes entry to today
def resize_days_table(oldest: int, workbook):


    oldest = datetime.datetime(year=oldest, month=1, day=1)
    today = datetime.datetime.today()
    df_days = pd.DataFrame()
    df_days["dates"] = pd.date_range(start=oldest,end=today)

    df_days["aux"] = df_days.reset_index().index + 2
    

    df_days["hours"] = df_days.agg(lambda x: f'''=SUMIFS(tb_list[Hours a Day],tb_list[Start Date],"<="&A{x['aux']},
                                    tb_list[Finish Date],">="&A{x['aux']})''', axis=1)
    df_days["ep"] = df_days.agg(lambda x: f'''=SUMIFS(tb_list[EP a Day],tb_list[Start Date],"<="&A{x['aux']},
                                    tb_list[Finish Date],">="&A{x['aux']})''', axis=1)

    #season shows is considered shows from the current and last year
    df_days["hours season"] = df_days.agg(lambda x: f'''=SUMIFS(tb_list[Hours a Day],tb_list[Start Date],
                                        "<="&A{x['aux']},tb_list[Finish Date],">="&A{x['aux']},tb_list[Year],
                                        days!G{x['aux']})+SUMIFS(tb_list[Hours a Day],tb_list[Start Date],
                                        "<="&A{x['aux']},tb_list[Finish Date],">="&A{x['aux']},tb_list[Year],
                                        days!G{x['aux']}-1)''', axis=1)
    df_days["finished shows"] = df_days.agg(lambda x: f'''=COUNTIF(user_list!G:G,A{x['aux']})''', axis=1)
    df_days["month"] = df_days.agg(lambda x: f'''=MONTH(A{x['aux']})''', axis=1)
    df_days["year"] = df_days.agg(lambda x: f'''=YEAR(A{x['aux']})''', axis=1)

    df_days.drop(columns=["aux"], inplace=True)


    return df_days

#TODO maybe do like in resize_days_table and add formula to dataframe instead of iterating over everything
# updates table size to fit the new entries
def resize_table_progression(workbook, oldest: int):
    sheet = workbook['progression']
    table = sheet.tables['tb_progression_month']
    #setting border style
    border = Border(bottom=Side(style='medium'))

    # RESIZE THE MONTH TABLE
    year = oldest
    today_year = datetime.datetime.now().year
    row = 3
    cell = "B3"
    sheet[cell] = oldest
    cell = "C3"
    datetime_month = datetime.datetime.strptime("1", "%m")
    sheet[cell] = datetime_month.strftime("%b")
    row = row + 1

    for i in range(2, 13):
        # add border when year change
        if i == 12:
            tb_range = sheet['B{}'.format(row):'I{}'.format(row)]
            for cell in tb_range:
                for x in cell:
                    x.border = border

        cellyear = "B{}".format(row)
        cellmonth = "C{}".format(row)
        cellmonthnumber = "D{}".format(row)
        cellep = "E{}".format(row)
        cellhours = "F{}".format(row)
        cellfinish = "G{}".format(row)
        cellpercentage = "H{}".format(row)
        cellhoursday = "I{}".format(row)

        sheet[cellyear] = '=B{}'.format(row - 1)
        datetime_moth = datetime.datetime.strptime(str(i), "%m")
        sheet[cellmonth] = datetime_moth.strftime("%b")
        sheet[cellmonthnumber] = i
        #episodes watched
        sheet[cellep] = "=SUMIFS(days!C:C,days!G:G,B{},days!F:F,D{})".format(row, row)
        #hours watched
        sheet[cellhours] = "=SUMIFS(days!B:B,days!G:G,B{},days!F:F,D{})".format(row, row)
        #series finished
        sheet[cellfinish] = "=SUMIFS(days!E:E,days!F:F,progression!D{},days!G:G,progression!B{})".format(
            row, row)
        #percentage of sesonals shows
        sheet[cellpercentage] = "=IF(F{}>0,SUMIFS(days!D:D,days!G:G,B{},days!F:F,D{})/F{},0)".format(row, row, row, row)
        sheet[cellhoursday] = "=F{}/DAY(EOMONTH(DATE(B{},D{},1),0))".format(row, row, row)
        row = row + 1
    year = year +1


    #the loops are separate because the first row of the table must have the value hardcoded
    #the first year loop if from 2 to 12 and the following are from the 1 to 12
    if oldest < today_year:
        while year <= today_year:
            for i in range(1, 13):
                # add border when year change
                if i == 12:
                    tb_range = sheet['B{}'.format(row):'I{}'.format(row)]
                    for cell in tb_range:
                        for x in cell:
                            x.border = border

                cellyear = "B{}".format(row)
                cellmonth = "C{}".format(row)
                cellmonthnumber = "D{}".format(row)
                cellep = "E{}".format(row)
                cellhours = "F{}".format(row)
                cellfinish = "G{}".format(row)
                cellpercentage = "H{}".format(row)
                cellhoursday = "I{}".format(row)

                sheet[cellyear] = '=B{}+1'.format(row - 12)
                datetime_moth = datetime.datetime.strptime(str(i), "%m")
                sheet[cellmonth] = datetime_moth.strftime("%b")
                sheet[cellmonthnumber] = i
                sheet[cellep] = "=SUMIFS(days!C:C,days!G:G,B{},days!F:F,D{})".format(row, row)
                sheet[cellhours] = "=SUMIFS(days!B:B,days!G:G,B{},days!F:F,D{})".format(row, row)
                sheet[cellfinish] = "=SUMIFS(days!E:E,days!F:F,progression!D{},days!G:G,progression!B{})".format(
                    row, row)
                sheet[cellpercentage] = "=IF(F{}>0,SUMIFS(days!D:D,days!G:G,B{},days!F:F,D{})/F{},0)".format(row, row,
                                                                                                             row, row)
                sheet[cellhoursday] = "=F{}/DAY(EOMONTH(DATE(B{},D{},1),0))".format(row, row, row)
                row = row + 1
            year = year + 1
        table.ref = 'B2:I' + str(row - 1)



    # RESIZE THE SEASON TABLE
    seasonyear = oldest
    seasonrow = 3
    cellseasonyear = "L{}".format(seasonrow)
    sheet[cellseasonyear] = "=B3"
    cellseason = "M{}".format(seasonrow)
    sheet[cellseason] = "WINTER"
    cellseasonep = "N{}".format(seasonrow)
    sheet[cellseasonep] = '''=SUMIFS(tb_progression_month[ep],tb_progression_month[month2],
    IF(M{}="WINTER","<=3",IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),tb_progression_month[month2],
    IF(M{}="WINTER",">=1",IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),tb_progression_month[year],
    L{})'''.format(seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow)
    cellseasonhours = "O{}".format(seasonrow)
    sheet[cellseasonhours] = '''=SUMIFS(tb_progression_month[Hours Watched],tb_progression_month[month2],
    IF(M{}="WINTER","<=3",IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),tb_progression_month[month2],
    IF(M{}="WINTER",">=1",IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),tb_progression_month[year],
    L{})'''.format(seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow)
    cellseasonfinish = "P{}".format(seasonrow)
    sheet[cellseasonfinish] = '''=SUMIFS(tb_progression_month[finished shows],tb_progression_month[month2],
    IF(M{}="WINTER","<=3",IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),tb_progression_month[month2],
    IF(M{}="WINTER",">=1",IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),tb_progression_month[year],
    L{})'''.format(seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow)
    cellseasonpercentage = "Q{}".format(seasonrow)
    sheet[cellseasonpercentage] = '''=IF(O{}>0,SUMIFS(days!D:D,days!F:F,IF(M{}="WINTER","<=3",
    IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),days!F:F,IF(M{}="WINTER",">=1",
    IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),days!G:G,L{})/O{},0)'''.format(seasonrow, seasonrow,
                                                                                         seasonrow, seasonrow,
                                                                                         seasonrow, seasonrow,
                                                                                         seasonrow, seasonrow,
                                                                                         seasonrow)
    cellseasonhoursday = "R{}".format(seasonrow)
    # I got lazy and considered each month month with 30 days
    sheet[cellseasonhoursday] = '''=SUMIFS(tb_progression_month[Hours Watched],tb_progression_month[month2],
    IF(M{}="WINTER","<=3",IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),tb_progression_month[month2],
    IF(M{}="WINTER",">=1",IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),tb_progression_month[year],L{})
    /90'''.format(seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow)
    seasonrow = seasonrow+1

    for i in range(2, 5):

        cellseason = "M{}".format(seasonrow)
        if i == 2:
            sheet[cellseason] = "SPRING"
        elif i == 3:
            sheet[cellseason] = "SUMMER"
        elif i == 4:
            sheet[cellseason] = "FALL"
            tb_range = sheet['L{}'.format(seasonrow):'R{}'.format(seasonrow)]
            # add border when year change
            for cell in tb_range:
                for x in cell:
                    x.border = border

        cellseasonep = "N{}".format(seasonrow)
        sheet[cellseasonep] = '''=SUMIFS(tb_progression_month[ep],tb_progression_month[month2],
        IF(M{}="WINTER","<=3",IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),tb_progression_month[month2],
        IF(M{}="WINTER",">=1",IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),tb_progression_month[year],
        L{})'''.format(seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow-i+1)
        cellseasonhours = "O{}".format(seasonrow)
        sheet[cellseasonhours] = '''=SUMIFS(tb_progression_month[Hours Watched],tb_progression_month[month2],
        IF(M{}="WINTER","<=3",IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),tb_progression_month[month2],
        IF(M{}="WINTER",">=1",IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),tb_progression_month[year],
        L{})'''.format(seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow-i+1)
        cellseasonfinish = "P{}".format(seasonrow)
        sheet[cellseasonfinish] = '''=SUMIFS(tb_progression_month[finished shows],tb_progression_month[month2],
        IF(M{}="WINTER","<=3",IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),tb_progression_month[month2],
        IF(M{}="WINTER",">=1",IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),tb_progression_month[year],
        L{})'''.format(seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow-i+1)
        cellseasonpercentage = "Q{}".format(seasonrow)
        sheet[cellseasonpercentage] = '''=IF(O{}>0,SUMIFS(days!D:D,days!F:F,IF(M{}="WINTER","<=3",
        IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),days!F:F,IF(M{}="WINTER",">=1",
        IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),days!G:G,L{})/O{},0)'''.format(seasonrow, seasonrow,
                                                                                             seasonrow, seasonrow,
                                                                                             seasonrow, seasonrow,
                                                                                             seasonrow, seasonrow-i+1,
                                                                                             seasonrow)
        cellseasonhoursday = "R{}".format(seasonrow)
        # I got lazy and considered each month month with 30 days
        sheet[cellseasonhoursday] = '''=SUMIFS(tb_progression_month[Hours Watched],tb_progression_month[month2],
        IF(M{}="WINTER","<=3",IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),tb_progression_month[month2],
        IF(M{}="WINTER",">=1",IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),tb_progression_month[year],L{})
        /90'''.format(seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow-i+1)
        seasonrow = seasonrow + 1
        



    seasonyear = seasonyear + 1
    if oldest <= today_year:
        while seasonyear <= today_year:

            for i in range(1,5):


                cellseason = "M{}".format(seasonrow)
                if i == 1:
                    sheet[cellseason] = "WINTER"
                    cellseasonyear = "L{}".format(seasonrow)
                    sheet[cellseasonyear] = "=L{}+1".format(seasonrow - 4)
                elif i ==2:
                    sheet[cellseason] = "SPRING"
                elif i == 3:
                    sheet[cellseason] = "SUMMER"
                elif i == 4:
                    sheet[cellseason] = "FALL"
                    tb_range = sheet['L{}'.format(seasonrow):'R{}'.format(seasonrow)]
                    # add border when year change
                    for cell in tb_range:
                        for x in cell:
                            x.border = border

                cellseasonep = "N{}".format(seasonrow)
                sheet[cellseasonep] = '''=SUMIFS(tb_progression_month[ep],tb_progression_month[month2],
                        IF(M{}="WINTER","<=3",IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),tb_progression_month[month2],
                        IF(M{}="WINTER",">=1",IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),tb_progression_month[year],
                        L{})'''.format(seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow-i+1)
                cellseasonhours = "O{}".format(seasonrow)
                sheet[cellseasonhours] = '''=SUMIFS(tb_progression_month[Hours Watched],tb_progression_month[month2],
                        IF(M{}="WINTER","<=3",IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),tb_progression_month[month2],
                        IF(M{}="WINTER",">=1",IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),tb_progression_month[year],
                        L{})'''.format(seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow-i+1)
                cellseasonfinish = "P{}".format(seasonrow)
                sheet[cellseasonfinish] = '''=SUMIFS(tb_progression_month[finished shows],tb_progression_month[month2],
                        IF(M{}="WINTER","<=3",IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),tb_progression_month[month2],
                        IF(M{}="WINTER",">=1",IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),tb_progression_month[year],
                        L{})'''.format(seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow-i+1)
                cellseasonpercentage = "Q{}".format(seasonrow)
                sheet[cellseasonpercentage] = '''=IF(O{}>0,SUMIFS(days!D:D,days!F:F,IF(M{}="WINTER","<=3",
                        IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),days!F:F,IF(M{}="WINTER",">=1",
                        IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),days!G:G,L{})/O{},
                        0)'''.format(seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow,
                                     seasonrow-i+1, seasonrow)
                cellseasonhoursday = "R{}".format(seasonrow)
                # I got lazy and considered each month month with 30 days
                sheet[cellseasonhoursday] = '''=SUMIFS(tb_progression_month[Hours Watched],tb_progression_month[month2],
                        IF(M{}="WINTER","<=3",IF(M{}="SPRING","<=6",IF(M{}="SUMMER","<=9","<=12"))),tb_progression_month[month2],
                        IF(M{}="WINTER",">=1",IF(M{}="SPRING",">=4",IF(M{}="SUMMER",">=7",">=10"))),tb_progression_month[year],L{})
                        /90'''.format(seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow, seasonrow-i+1)

                seasonrow = seasonrow +1
            seasonyear = seasonyear + 1

    sheet.tables['tb_progression_season'].ref = 'L2:R' + str(seasonrow - 1)

    # i had to add the conditional formatting in the code because even though the table size is increase the conditional
    # formatting remains the same
    # Note to pastself - if you enter the entire column in the conditional formatting you don't need this but i'll leave it here
    # for the month table the season has de conditional formatting in the entire column there's no need

    sheet.conditional_formatting.add("E3:E{}".format(row - 1),
                                     ColorScaleRule(start_type='min', start_color='f8696b',
                                                    mid_type='percentile', mid_value=50, mid_color='ffeb84',
                                                    end_type='max', end_color='63be7b'))
    sheet.conditional_formatting.add("F3:F{}".format(row - 1),
                                     ColorScaleRule(start_type='min', start_color='f8696b',
                                                    mid_type='percentile', mid_value=50, mid_color='ffeb84',
                                                    end_type='max', end_color='63be7b'))


    sheet.conditional_formatting.add("H3:H{}".format(row - 1),
                                     DataBarRule(start_type='num', start_value=0, end_type='num', end_value='1',
                                                 color="FF638EC6", showValue="None", minLength=None, maxLength=None))
    sheet.conditional_formatting.add("I3:I{}".format(row - 1),
                                     IconSetRule('3Symbols', 'percent', [0, 33, 67], showValue=None, percent=None))

    return workbook

#I had this table made in excel by openpyxl messed up my array formulas
#(I tried open and saving the file without doing anything and it still mess my formulas)
def resize_genres(workbook, genreslist):
    sheet = workbook['challenge']
    row = 3
    for genre in genreslist:
        cellgenres = "J{}".format(row)
        sheet[cellgenres] = genre
        celltotal = "K{}".format(row)
        sheet[celltotal] = '''=COUNTIFS(tb_anime_genres[Genres],J{}, tb_anime_genres[Start Date],
        ">=" & $B$1,tb_anime_genres[Finish Date],"<="&$B$2)'''.format(row)
        cellpercentage = "L{}".format(row)
        sheet[cellpercentage] = '=K{}/$C$7'.format(row)
        #change sell format
        sheet[cellpercentage].number_format = "0.00%"

        cellaverage = "M{}".format(row)
        sheet[cellaverage] = '''=IF(K{}=0,0,AVERAGEIFS(tb_anime_genres[Score],tb_anime_genres[Genres],J{}, 
        tb_anime_genres[Start Date],">=" & $B$1,tb_anime_genres[Finish Date],"<="&$B$2))'''.format(row, row)
        cellhours = "N{}".format(row)
        sheet[cellhours] = '''=SUMIFS(tb_anime_genres[Show Duration (hours)],tb_anime_genres[GENRES],J{}, tb_anime_genres[Start Date],
        ">=" & $B$1,tb_anime_genres[Finish Date],"<="&$B$2)'''.format(row)
        cellweightedmean = "O{}".format(row)
        sheet[cellweightedmean] = '''=IF(K{}=0,0,M{}*(K{}/$C$7)+$C$15*($C$7-K{})/$C$7)'''.format(row, row, row, row) 
        cellweightedmean2 = "P{}".format(row)
        sheet[cellweightedmean2] = '''=IF(K{}=0,0,M{}*(K{}/(K{}+15))+$C$15*15/(K{}+15))'''.format(row, row, row, row, row)
        row = row + 1

    sheet.tables['tb_challenge'].ref = 'I2:P' + str(row - 1)

    #basically the same thing as the previous but overall instead ofbeing for values between dates
    sheet = workbook['genres_stats']
    row = 2
    for genre in genreslist:
        cellgenres = "A{}".format(row)
        sheet[cellgenres] = genre
        celltotal = "B{}".format(row)
        sheet[celltotal] = '''=COUNTIF(tb_anime_genres[Genres],A{})'''.format(row)
        cellpercentage = "C{}".format(row)
        sheet[cellpercentage] = '=B{}/COUNTIF(tb_list[Title],"<>"&"")'.format(row)
        sheet[cellpercentage].number_format="0.00%"
        cellaverage = "D{}".format(row)
        sheet[cellaverage] = '''=AVERAGEIFS(tb_anime_genres[Score],tb_anime_genres[Genres],A{})'''.format(row)
        cellhours = "E{}".format(row)
        sheet[cellhours] = '''=SUMIFS(tb_anime_genres[Show Duration (hours)],tb_anime_genres[Genres],A{})'''.format(row)
        cellweightedmean = "F{}".format(row)
        sheet[cellweightedmean] = '''=D{}*(B{}/COUNTIF(tb_list[Title], "<>"&""))+
        AVERAGEIF(tb_list[Score],"<>"&0,tb_list[Score])*(COUNTIF(tb_list[Title],"<>"&"")-
        B{})/COUNTIF(tb_list[Title],"<>"&"")'''.format(row, row, row)
        cellweightedmean2 = "G{}".format(row)
        sheet[cellweightedmean2] = '''=D{}*COUNTIF(tb_list[Title], "<>"&"")/(COUNTIF(tb_list[Title], "<>"&"")+15)+
          AVERAGEIF(tb_list[Score],"<>"&0,tb_list[Score])*15/(COUNTIF(tb_list[Title], "<>"&"")+15)'''.format(row)
        row = row + 1

    sheet.tables['tb_genres'].ref = 'A1:G' + str(row - 1)

    return workbook

#TODO this is wrong i need to evaluate the seasonality and then choose the aproppriate method
def forecast_table(workbook,oldest):
    sheet = workbook['forecast -amort exp']
    row = 4
    cellmt="F8{}".format(row)
    sheet[cellmt] = "=AVERAGE(progression!O3:O6)"

    #setting value for the first year recorded
    for i in range(1,5):
        row = row + 1
        cellyear="B{}".format(row)
        sheet[cellyear] = oldest
        cellseason="C{}".format(row)
        sheet[cellseason] = i
        cellzt="D{}".format(row)
        sheet[cellzt] = "=progression!O{}".format((row-2))
        cellst="I{}".format(row)
        sheet[cellst] = "=D{}-$F$8".format(row)

    # adding following years until the year previous to the corrent one
    oldest = oldest + 1
    print(oldest)
    while oldest < datetime.date.today().year:
        for i in range(1,5):
            row = row + 1
            cellyear="B{}".format(row)
            sheet[cellyear] = "=IF(C{}=1,B{}+1,B{})".format(row,row-1,row-1)
            cellseason="C{}".format(row)
            sheet[cellseason] = "=C{}".format(row-4)
            cellzt="D{}".format(row)
            sheet[cellzt] = "=progression!O{}".format((row-2))
            cellzt2="E{}".format(row)
            sheet[cellzt2] = "=D{}-I{}".format(row,row-4)
            cellmt="F{}".format(row)
            sheet[cellmt] = "=+$D$1*E{}+(1-$D$1)*F{}".format(row, row-1)
            cellyt2="G{}".format(row)
            sheet[cellyt2] = "=F{}".format(row-1)
            cellyt="H{}".format(row)
            sheet[cellyt] = "=G{}+I{}".format(row, row-4)
            cellst="I{}".format(row)
            sheet[cellst] = "=+$D$2*(D{}-F{})+(1-$D$2)*I{}".format(row,row,row-4)
            cellzy="J{}".format(row)
            sheet[cellzy] = "=+ABS(D{}-H{})".format(row,row)
            cella="k{}".format(row)
            sheet[cella] = "=+$D$1*J{}+(1-$D$1)*K{}".format(row, row-1)
            celle="L{}".format(row)
            sheet[celle] = "=SQRT(PI()/2)*K{}".format(row)
            cellc="M{}".format(row)
            sheet[cellc] = "=H{}+2*L{}".format(row,row)
            cellb="N{}".format(row)
            sheet[cellb] = "=H{}-2*L{}".format(row,row)
            cellprev="O{}".format(row)
            sheet[cellprev] = '''=IF(AND(D{}<=M{},D{}>=N{}),"ok","not ok")'''.format(row,row,row,row)
            celltest="A{}".format(row)
            sheet[celltest] = 'feito'
        oldest = oldest + 1 

    #check current season to stop in the previous one
    if datetime.date.today().month <=3:
        season = 1
    elif datetime.date.today().month <=6:
        season = 2
    elif datetime.date.today().month <=9:
        season = 3
    elif datetime.date.today().month <=12:
        season = 4

    for i in range(1,season):
            #TODO: this code is the same as in the privous it might be better to make a function to avoid repeating it
            row = row + 1
            cellyear="B{}".format(row)
            sheet[cellyear] = "=IF(C{}=1,B{}+1,B{})".format(row,row-1,row-1)
            cellseason="C{}".format(row)
            sheet[cellseason] = "=C{}".format(row-4)
            cellzt="D{}".format(row)
            sheet[cellzt] = "=progression!O{}".format((row-2))
            cellzt2="E{}".format(row)
            sheet[cellzt2] = "=D{}-I{}".format(row,row-4)
            cellmt="F{}".format(row)
            sheet[cellmt] = "=+$D$1*E{}+(1-$D$1)*F{}".format(row, row-1)
            cellyt2="G{}".format(row)
            sheet[cellyt2] = "=F{}".format(row-1)
            cellyt="H{}".format(row)
            sheet[cellyt] = "=G{}+I{}".format(row, row-4)
            cellst="I{}".format(row)
            sheet[cellst] = "=+$D$2*(D{}-F{})+(1-$D$2)*I{}".format(row,row,row-4)
            cellzy="J{}".format(row)
            sheet[cellzy] = "=+ABS(D{}-H{})".format(row,row)
            cella="k{}".format(row)
            sheet[cella] = "=+$D$1*J{}+(1-$D$1)*K{}".format(row, row-1)
            celle="L{}".format(row)
            sheet[celle] = "=SQRT(PI()/2)*K{}".format(row)
            cellc="M{}".format(row)
            sheet[cellc] = "=H{}+2*L{}".format(row,row)
            cellb="N{}".format(row)
            sheet[cellb] = "=H{}-2*L{}".format(row,row)
            cellprev="O{}".format(row)
            sheet[cellprev] = '''=IF(AND(D{}<=M{},D{}>=N{}),"ok","not ok")'''.format(row,row,row,row)
            celltest="A{}".format(row)
            sheet[celltest] = 'feito'


    ##previous for the following year
    fixrow = row +1 
    for i in range(1,4):
            row = row + 1

            cellyear="B{}".format(row)
            sheet[cellyear] = "=IF(C{}=1,B{}+1,B{})".format(row,row-1,row-1)
            cellseason="C{}".format(row)
            sheet[cellseason] = "=C{}".format(row-4)
            cellyt2="G{}".format(row)
            sheet[cellyt2] = "=F{}".format(fixrow-1)
            cellyt="H{}".format(row)
            sheet[cellyt] = "=G{}+I{}".format(row, row-4)

    ##add border to table
    numberyears=int(np.ceil((row-4)/4))
    print(numberyears)
    border = Border(bottom=Side(style='medium'))
    for i in range(1, numberyears):
        tb_range = sheet['B{}'.format(4+i*4):'k{}'.format(4+i*4)]
        for cell in tb_range:
            for x in cell:
                x.border = border

    return workbook


